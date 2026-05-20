# app.py — Smart Remarks Tool v2.0
# =====================================================================
import sys, io, re, logging, unicodedata, json, hashlib, base64
from datetime import datetime
import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.chart import BarChart, LineChart, RadarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots

# =====================================================================
# 1. التكوين الشامل (Master Configuration)
# =====================================================================

MASTER_CONFIG = {
    "APP_VERSION": "2.0.0",
    "APP_NAME": "أداة الملاحظات الذكية",
    "SUBJECT_PATTERNS": {
        "arabic": {
            "keywords": ["عربي", "العربية", "لغة عربية", "ع.ع", "ل.ع"],
            "lang": "ar", "icon": "📖", "color": "#16A085",
            "label": "اللغة العربية",
            "weight": 3
        },
        "math": {
            "keywords": ["رياض", "رياضيات", "math", "ر.ع", "ريا"],
            "lang": "ar", "icon": "🔢", "color": "#8E44AD",
            "label": "الرياضيات",
            "weight": 3
        },
        "french": {
            "keywords": ["فرنس", "الفرنسية", "français", "francais", "ف.ع", "ل.ف"],
            "lang": "fr", "icon": "🇫🇷", "color": "#2980B9",
            "label": "اللغة الفرنسية",
            "weight": 2
        },
        "english": {
            "keywords": ["انجليز", "english", "ل.ا", "eng"],
            "lang": "en", "icon": "🇬🇧", "color": "#1ABC9C",
            "label": "اللغة الإنجليزية",
            "weight": 2
        },
        "tamazight": {
            "keywords": ["أمازيغ", "امازيغ", "tamazight", "ل.أ"],
            "lang": "ar", "icon": "ⵣ", "color": "#E67E22",
            "label": "الأمازيغية",
            "weight": 1
        },
        "islamic": {
            "keywords": ["تربية إسلامية", "إسلامية", "دين", "قرآن", "إسلام"],
            "lang": "ar", "icon": "☪️", "color": "#27AE60",
            "label": "التربية الإسلامية",
            "weight": 2
        },
        "civic": {
            "keywords": ["تربية مدنية", "مدنية", "مواطنة"],
            "lang": "ar", "icon": "🏛️", "color": "#E74C3C",
            "label": "التربية المدنية",
            "weight": 1
        },
        "science": {
            "keywords": ["علوم", "طبيعة", "بيئة", "علوم التكنولوجيا"],
            "lang": "ar", "icon": "🔬", "color": "#3498DB",
            "label": "العلوم التكنولوجيا",
            "weight": 2
        },
        "history_geo": {
            "keywords": ["تاريخ", "جغرافيا", "تاريخ وجغرافيا"],
            "lang": "ar", "icon": "🌍", "color": "#95A5A6",
            "label": "التاريخ والجغرافيا",
            "weight": 1
        },
        "plastic_arts": {
            "keywords": ["تشكيل", "تشكيلية", "رسم", "art plastique"],
            "lang": "ar", "icon": "🎨", "color": "#F39C12",
            "label": "التربية التشكيلية",
            "weight": 1
        },
        "music": {
            "keywords": ["موسيق", "موسيقية", "نشيد", "musique"],
            "lang": "ar", "icon": "🎵", "color": "#9B59B6",
            "label": "التربية الموسيقية",
            "weight": 1
        },
        "physical_ed": {
            "keywords": ["بدنية", "رياضية", "sport", "eps"],
            "lang": "ar", "icon": "⚽", "color": "#E74C3C",
            "label": "التربية البدنية",
            "weight": 1
        },
    },
    "COL_PATTERNS": {
        "base_headers": ["الاسم", "اللقب", "الرقم", "رقم التعريف", "التلميذ", "المعدل"],
        "name_col": ["الاسم واللقب", "الاسم", "اللقب", "التلميذ", "nom", "prénom"],
        "ar_expr": ["تعبير", "تواصل شفوي", "شفوي", "التعبير"],
        "ar_read": ["قراءة", "محفوظات", "القراءة"],
        "ar_write": ["كتابة", "إملاء", "الكتابة"],
        "ma_num": ["أعداد", "عمليات", "حساب", "الأعداد"],
        "ma_meas": ["مقادير", "قياس", "القياس"],
        "ma_data": ["تنظيم", "معطيات", "إحصاء"],
        "ma_geo": ["فضاء", "هندسة", "الهندسة"],
        "fl_expr": ["expression", "orale", "شفهي", "تواصل"],
        "fl_read": ["lecture", "compréhension", "قراءة"],
        "fl_prod": ["production", "écrite", "إنتاج", "كتابي"],
        "plastic_eval": ["تقويم", "إنجاز", "عمل", "مشاركة", "علامة"],
        "music_eval": ["نشيد", "أداء", "استماع", "إيقاع", "علامة"],
        "sport_eval": ["أداء", "نشاط", "مهارة", "علامة"],
        "exam": ["اختبار", "فرض", "امتحان", "exam", "test"],
        "remark": ["ملاحظة", "تقدير", "remarque", "appreciation", "تعليق"]
    },
    "REMARKS": {
        "ar": [
            (0.95, "ممتاز +"),
            (0.85, "ممتاز"),
            (0.75, "جيد جداً"),
            (0.65, "جيد"),
            (0.55, "مقبول"),
            (0.50, "متوسط"),
            (0.00, "دون المتوسط")
        ],
        "fr": [
            (0.95, "Excellent +"),
            (0.85, "Excellent"),
            (0.75, "Très bien"),
            (0.65, "Bien"),
            (0.55, "Passable"),
            (0.50, "Suffisant"),
            (0.00, "Insuffisant")
        ],
        "en": [
            (0.95, "Outstanding +"),
            (0.85, "Excellent"),
            (0.75, "Very Good"),
            (0.65, "Good"),
            (0.55, "Acceptable"),
            (0.50, "Sufficient"),
            (0.00, "Insufficient")
        ],
    },
    "GRADE_COLORS": {
        "gradient": {
            (0.95, 1.01): "#1A6B3C",
            (0.85, 0.95): "#27AE60",
            (0.75, 0.85): "#52BE80",
            (0.65, 0.75): "#F1C40F",
            (0.55, 0.65): "#E67E22",
            (0.50, 0.55): "#E74C3C",
            (0.00, 0.50): "#C0392B",
        }
    },
    "COLORS": {
        "ar": {
            "ممتاز +": "1A6B3C", "ممتاز": "27AE60",
            "جيد جداً": "52BE80", "جيد": "F1C40F",
            "مقبول": "E67E22", "متوسط": "E74C3C",
            "دون المتوسط": "C0392B",
            "غائب": "95A5A6", "معفى": "7F8C8D"
        },
        "fr": {
            "Excellent +": "1A6B3C", "Excellent": "27AE60",
            "Très bien": "52BE80", "Bien": "F1C40F",
            "Passable": "E67E22", "Suffisant": "E74C3C",
            "Insuffisant": "C0392B",
            "Absent": "95A5A6", "Dispensé": "7F8C8D"
        },
        "en": {
            "Outstanding +": "1A6B3C", "Excellent": "27AE60",
            "Very Good": "52BE80", "Good": "F1C40F",
            "Acceptable": "E67E22", "Sufficient": "E74C3C",
            "Insufficient": "C0392B",
            "Absent": "95A5A6", "Exempt": "7F8C8D"
        }
    },
    # ========= إعدادات التصنيف والشهادة ==========
    "RANKS": {
        "ar": ["الأول", "الثاني", "الثالث", "الرابع", "الخامس",
               "السادس", "السابع", "الثامن", "التاسع", "العاشر"],
        "fr": ["1er", "2ème", "3ème", "4ème", "5ème",
               "6ème", "7ème", "8ème", "9ème", "10ème"],
    },
    # ========= الفصول الدراسية ====================
    "TRIMESTERS": ["الفصل الأول", "الفصل الثاني", "الفصل الثالث"],
    "LEVELS": [
        "السنة الأولى ابتدائي", "السنة الثانية ابتدائي",
        "السنة الثالثة ابتدائي", "السنة الرابعة ابتدائي",
        "السنة الخامسة ابتدائي"
    ],
}

# =====================================================================
# 2. النظام المساعد (Helper Layer)
# =====================================================================

class TextHelper:
    @staticmethod
    def normalize(text: str) -> str:
        if not text: return ""
        text = str(text).strip().lower()
        text = unicodedata.normalize("NFD", text)
        text = "".join(c for c in text if unicodedata.category(c) != "Mn")
        return re.sub(r"\s+", " ", text)

    @staticmethod
    def arabic_ordinal(n: int) -> str:
        words = MASTER_CONFIG["RANKS"]["ar"]
        if 1 <= n <= len(words): return words[n - 1]
        return str(n)

    @staticmethod
    def to_arabic_numerals(n) -> str:
        eastern = str.maketrans('0123456789', '٠١٢٣٤٥٦٧٨٩')
        return str(n).translate(eastern)


class GradeCalculator:
    def __init__(self, config):
        self.remarks = config["REMARKS"]
        self.grade_colors = config["GRADE_COLORS"]["gradient"]

    def parse_grade(self, value) -> dict:
        if value is None or str(value).strip() == "":
            return {"status": "empty", "value": None}
        v_str = str(value).strip().lower()
        if v_str in ["غ", "غائب", "abs", "absent"]:
            return {"status": "absent", "value": "غائب"}
        if v_str in ["م", "معفى", "dispensé", "dispense", "exempt", "مُعفى"]:
            return {"status": "exempt", "value": "معفى"}
        try:
            num = float(v_str.replace(",", "."))
            if num != num: return {"status": "invalid", "value": None}
            return {"status": "valid", "value": num}
        except ValueError:
            return {"status": "text", "value": v_str}

    def detect_max_grade(self, header_val, default: float) -> float:
        if not header_val: return default
        m = re.search(r"/\s*(\d+(?:\.\d+)?)", str(header_val))
        if m and float(m.group(1)) > 0: return float(m.group(1))
        if "20" in str(header_val): return 20.0
        if "10" in str(header_val): return 10.0
        return default

    def calc_average(self, values: list, ignore_zero: bool) -> float:
        valid = [v for v in values if v is not None]
        if ignore_zero: valid = [v for v in valid if v != 0.0]
        return sum(valid) / len(valid) if valid else None

    def get_remark(self, grade: float, max_grade: float, lang: str) -> str:
        if max_grade <= 0: return ""
        grade = max(0.0, min(float(grade), max_grade))
        ratio = grade / max_grade
        table = self.remarks.get(lang, self.remarks["ar"])
        for threshold, text in table:
            if ratio >= threshold: return text
        return table[-1][1]

    def get_grade_hex_color(self, grade: float, max_grade: float) -> str:
        if max_grade <= 0: return "95A5A6"
        ratio = grade / max_grade
        for (lo, hi), color in self.grade_colors.items():
            if lo <= ratio < hi:
                return color.lstrip("#")
        return "C0392B"

    def calc_statistics(self, grades: list, max_grade: float) -> dict:
        if not grades:
            return {}
        arr = np.array(grades)
        passing = [g for g in grades if g >= max_grade / 2]
        return {
            "count": len(grades),
            "mean": round(float(arr.mean()), 2),
            "median": round(float(np.median(arr)), 2),
            "std": round(float(arr.std()), 2),
            "min": round(float(arr.min()), 2),
            "max": round(float(arr.max()), 2),
            "pass_rate": round(len(passing) / len(grades) * 100, 1),
            "q1": round(float(np.percentile(arr, 25)), 2),
            "q3": round(float(np.percentile(arr, 75)), 2),
            "distribution": {
                "ممتاز +": sum(1 for g in grades if g / max_grade >= 0.95),
                "ممتاز": sum(1 for g in grades if 0.85 <= g / max_grade < 0.95),
                "جيد جداً": sum(1 for g in grades if 0.75 <= g / max_grade < 0.85),
                "جيد": sum(1 for g in grades if 0.65 <= g / max_grade < 0.75),
                "مقبول": sum(1 for g in grades if 0.55 <= g / max_grade < 0.65),
                "متوسط": sum(1 for g in grades if 0.50 <= g / max_grade < 0.55),
                "دون المتوسط": sum(1 for g in grades if g / max_grade < 0.50),
            }
        }


# =====================================================================
# 3. منشئ الكارنيه (Bulletin Generator)
# =====================================================================

class BulletinGenerator:
    """ينشئ كارنيه تلاميذ كاملة داخل ملف إكسيل منفصل"""

    def __init__(self, config, settings):
        self.config = config
        self.settings = settings
        self.calc = GradeCalculator(config)

    def create_bulletin_sheet(self, wb, student_data: dict,
                              school_info: dict, trimester: str):
        name = student_data.get("name", "تلميذ")
        safe = re.sub(r'[\\/*?:\[\]]', '_', name)[:28]
        ws = wb.create_sheet(title=f"كارنيه_{safe}")
        ws.sheet_view.rightToLeft = True
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 22

        thin = Side(style="thin", color="CCCCCC")
        med  = Side(style="medium", color="2C3E50")
        border_full = Border(left=med, right=med, top=med, bottom=med)
        border_thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))

        # -------- رأس الكارنيه --------
        ws.merge_cells("B1:F1")
        hdr = ws["B1"]
        hdr.value = school_info.get("ministry", "وزارة التربية الوطنية")
        hdr.font = Font(bold=True, size=11, color="1A252F")
        hdr.alignment = Alignment(horizontal="center", vertical="center")
        hdr.fill = PatternFill(fill_type="solid", fgColor="D6EAF8")
        ws.row_dimensions[1].height = 22

        ws.merge_cells("B2:F2")
        ws["B2"].value = (f"مديرية التربية لولاية: {school_info.get('wilaya', '')}   "
                          f"—  ابتدائية: {school_info.get('school', '')}")
        ws["B2"].font = Font(bold=True, size=10, color="2C3E50")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        ws.merge_cells("B3:F3")
        ws["B3"].value = f"كشف النقاط — {trimester}"
        ws["B3"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["B3"].fill = PatternFill(fill_type="solid", fgColor="2980B9")
        ws["B3"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 28

        # -------- بيانات التلميذ --------
        ws.row_dimensions[4].height = 8
        ws.merge_cells("B5:F5")
        ws["B5"].value = (f"الاسم واللقب: {name}    "
                          f"القسم: {student_data.get('class', '')}    "
                          f"السنة الدراسية: {school_info.get('year', '')}")
        ws["B5"].font = Font(bold=True, size=11, color="1A252F")
        ws["B5"].fill = PatternFill(fill_type="solid", fgColor="EBF5FB")
        ws["B5"].alignment = Alignment(horizontal="right",
                                       vertical="center", wrap_text=True)
        ws["B5"].border = border_full
        ws.row_dimensions[5].height = 24

        # -------- رؤوس الجدول --------
        ws.row_dimensions[6].height = 8
        headers = ["المادة", "نقطة التقويم", "نقطة الاختبار",
                   "المعدل", "الملاحظة"]
        header_cols = ["B", "C", "D", "E", "F"]
        ws.row_dimensions[7].height = 26
        for col_l, h in zip(header_cols, headers):
            c = ws[f"{col_l}7"]
            c.value = h
            c.font = Font(bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill(fill_type="solid", fgColor="1A252F")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border_thin

        # -------- صفوف المواد --------
        row = 8
        weighted_sum, total_weight = 0.0, 0
        all_avgs = []

        for subj_key, subj_data in student_data.get("subjects", {}).items():
            subj_cfg = self.config["SUBJECT_PATTERNS"].get(subj_key, {})
            label = subj_cfg.get("label", subj_key)
            icon  = subj_cfg.get("icon", "📘")
            lang  = subj_cfg.get("lang", "ar")
            weight = subj_cfg.get("weight", 1)
            max_g = subj_data.get("max", self.settings["max_grade"])

            eval_avg = subj_data.get("eval_avg")
            exam_val = subj_data.get("exam")
            final    = subj_data.get("final")

            if final is None:
                vals = [v for v in [eval_avg, exam_val] if v is not None]
                final = sum(vals) / len(vals) if vals else None

            remark_txt = ""
            hex_col = "FFFFFF"
            if final is not None:
                remark_txt = self.calc.get_remark(final, max_g, lang)
                hex_col    = self.calc.get_grade_hex_color(final, max_g)
                weighted_sum += final * weight
                total_weight += weight
                all_avgs.append(final)

            ws.row_dimensions[row].height = 22
            ws[f"B{row}"].value = f"{icon} {label}"
            ws[f"C{row}"].value = (round(eval_avg, 2)
                                   if eval_avg is not None else "—")
            ws[f"D{row}"].value = (round(exam_val, 2)
                                   if exam_val is not None else "—")
            ws[f"E{row}"].value = (round(final, 2)
                                   if final is not None else "—")
            ws[f"F{row}"].value = remark_txt

            for col_l in ["B", "C", "D", "E", "F"]:
                c = ws[f"{col_l}{row}"]
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border_thin
                bg = "F8F9FA" if row % 2 == 0 else "FFFFFF"
                c.fill = PatternFill(fill_type="solid", fgColor=bg)

            ws[f"E{row}"].fill = PatternFill(fill_type="solid",
                                             fgColor=hex_col)
            ws[f"E{row}"].font = Font(bold=True, color="FFFFFF")
            ws[f"F{row}"].fill = PatternFill(fill_type="solid",
                                             fgColor=hex_col)
            ws[f"F{row}"].font = Font(bold=True, color="FFFFFF", size=9)
            row += 1

        # -------- المعدل العام --------
        ws.row_dimensions[row].height = 8
        row += 1
        gen_avg = (weighted_sum / total_weight) if total_weight else None
        gen_hex = (self.calc.get_grade_hex_color(gen_avg, self.settings["max_grade"])
                   if gen_avg else "95A5A6")
        gen_remark = (self.calc.get_remark(gen_avg, self.settings["max_grade"], "ar")
                      if gen_avg else "—")

        ws.merge_cells(f"B{row}:D{row}")
        ws[f"B{row}"].value = "🏆 المعدل العام المُرجَّح"
        ws[f"B{row}"].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f"B{row}"].fill = PatternFill(fill_type="solid", fgColor="1A252F")
        ws[f"B{row}"].alignment = Alignment(horizontal="center",
                                            vertical="center")
        ws.row_dimensions[row].height = 28

        ws[f"E{row}"].value = (round(gen_avg, 2) if gen_avg else "—")
        ws[f"E{row}"].font = Font(bold=True, size=13, color="FFFFFF")
        ws[f"E{row}"].fill = PatternFill(fill_type="solid", fgColor=gen_hex)
        ws[f"E{row}"].alignment = Alignment(horizontal="center",
                                            vertical="center")

        ws[f"F{row}"].value = gen_remark
        ws[f"F{row}"].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f"F{row}"].fill = PatternFill(fill_type="solid", fgColor=gen_hex)
        ws[f"F{row}"].alignment = Alignment(horizontal="center",
                                            vertical="center")

        for col_l in ["B", "C", "D", "E", "F"]:
            ws[f"{col_l}{row}"].border = border_full

        # -------- توقيعات --------
        sig_row = row + 2
        ws.row_dimensions[sig_row].height = 20
        for col_l, label in zip(["B", "D", "F"],
                                 ["توقيع الأستاذ(ة)", "توقيع المدير(ة)",
                                  "توقيع الولي"]):
            ws[f"{col_l}{sig_row}"].value = label
            ws[f"{col_l}{sig_row}"].font = Font(bold=True, size=9,
                                                 color="7F8C8D")
            ws[f"{col_l}{sig_row}"].alignment = Alignment(
                horizontal="center")
            ws[f"{col_l}{sig_row}"].border = Border(
                bottom=Side(style="medium", color="7F8C8D"))

        return ws, gen_avg


# =====================================================================
# 4. مُعالج الإكسيل الرئيسي (Excel Processor v2)
# =====================================================================

class ExcelProcessorV2:
    def __init__(self, config, user_settings, logger):
        self.config  = config
        self.settings = user_settings
        self.logger  = logger
        self.calc    = GradeCalculator(config)
        self.stats   = []
        self.all_student_data = {}   # {sheet: {row: grade}}
        self.processing_log  = []

    # ---------- تحديد نوع المادة ----------
    def detect_subject(self, sheet_name: str):
        norm = TextHelper.normalize(sheet_name)
        for stype, info in self.config["SUBJECT_PATTERNS"].items():
            for kw in info["keywords"]:
                if TextHelper.normalize(kw) in norm:
                    return stype, info["lang"]
        return "other", "ar"

    # ---------- تحديد الأعمدة ----------
    def find_column(self, ws, keywords, max_row=20):
        norm_kws = [TextHelper.normalize(kw) for kw in keywords]
        for row in ws.iter_rows(min_row=1,
                                max_row=min(max_row, ws.max_row or 1)):
            for cell in row:
                if cell.value:
                    nv = TextHelper.normalize(str(cell.value))
                    if any(kw in nv for kw in norm_kws):
                        return {"col": cell.column, "row": cell.row,
                                "header": str(cell.value)}
        return None

    def find_name_column(self, ws, header_row):
        patterns = self.config["COL_PATTERNS"]["name_col"]
        info = self.find_column(ws, patterns)
        if info: return info["col"]
        return 1

    def find_fallback_numeric_cols(self, ws, header_row, exclude_cols,
                                   max_grade):
        numeric_cols = []
        if not ws.max_row or ws.max_row <= header_row:
            return numeric_cols
        check_rows = range(header_row + 1,
                           min(header_row + 5, ws.max_row + 1))
        for col in range(1, (ws.max_column or 1) + 1):
            if col in exclude_cols: continue
            for r in check_rows:
                p = self.calc.parse_grade(
                    ws.cell(row=r, column=col).value)
                if p["status"] == "valid" and \
                   0 <= p["value"] <= max_grade * 1.1:
                    numeric_cols.append(col)
                    break
        return numeric_cols

    def map_columns(self, ws, subject_type, default_max):
        res = {
            "found": False, "remark_col": None, "exam_col": None,
            "eval_cols": [], "header_row": 1, "exam_header": "",
            "method": "none", "notes": "", "name_col": 1
        }

        base_info = self.find_column(
            ws, self.config["COL_PATTERNS"]["base_headers"])
        if base_info:
            res["header_row"] = base_info["row"]

        exam_info = self.find_column(
            ws, self.config["COL_PATTERNS"]["exam"])
        if exam_info:
            res["exam_col"] = exam_info["col"]
            res["exam_header"] = exam_info["header"]
            res["header_row"] = max(res["header_row"], exam_info["row"])

        grp_map = {
            "arabic":       ["ar_expr", "ar_read", "ar_write"],
            "math":         ["ma_num", "ma_meas", "ma_data", "ma_geo"],
            "french":       ["fl_expr", "fl_read", "fl_prod"],
            "english":      ["fl_expr", "fl_read", "fl_prod"],
            "tamazight":    ["fl_expr", "fl_read", "fl_prod"],
            "islamic":      ["plastic_eval"],
            "civic":        ["plastic_eval"],
            "science":      ["plastic_eval"],
            "history_geo":  ["plastic_eval"],
            "plastic_arts": ["plastic_eval"],
            "music":        ["music_eval"],
            "physical_ed":  ["sport_eval"],
        }
        for grp in grp_map.get(subject_type, []):
            info = self.find_column(
                ws, self.config["COL_PATTERNS"].get(grp, []))
            if info:
                res["eval_cols"].append(info["col"])
                res["header_row"] = max(res["header_row"], info["row"])

        rem_info = self.find_column(
            ws, self.config["COL_PATTERNS"]["remark"])
        if rem_info:
            res["remark_col"] = rem_info["col"]
            res["header_row"] = max(res["header_row"], rem_info["row"])
        else:
            res["remark_col"] = (ws.max_column or 0) + 1
            hc = ws.cell(row=res["header_row"],
                         column=res["remark_col"])
            hc.value = "ملاحظة الأستاذ"
            hc.font  = Font(bold=True, color="FFFFFF")
            hc.fill  = PatternFill(fill_type="solid", fgColor="2980B9")
            res["notes"] = "تم إنشاء عمود ملاحظات تلقائياً"

        res["name_col"] = self.find_name_column(ws, res["header_row"])
        num_cols = self.find_fallback_numeric_cols(
            ws, res["header_row"], {res["remark_col"]}, default_max)

        if num_cols and not res["eval_cols"] and res["exam_col"] is None:
            if len(num_cols) > 1:
                res["exam_col"]  = num_cols[-1]
                res["eval_cols"] = num_cols[:-1]
            else:
                res["eval_cols"] = num_cols
            res["method"] = "fallback"

        res["found"] = bool(
            num_cols or res["eval_cols"] or res["exam_col"] is not None)
        return res

    # ---------- حساب المعدل النهائي ----------
    def compute_final_grade(self, eval_vals, exam_val, subject_type,
                            weights=None):
        eval_avg = (self.calc.calc_average(eval_vals, self.settings["ignore_zero"])
                    if eval_vals else None)
        if subject_type in ("arabic", "math", "french", "english",
                            "tamazight", "islamic", "civic", "science",
                            "history_geo"):
            ew = weights.get("eval", 1) if weights else 1
            xw = weights.get("exam", 1) if weights else 1
            if eval_avg is not None and exam_val is not None:
                return (eval_avg * ew + exam_val * xw) / (ew + xw)
            if eval_avg is not None: return eval_avg
            if exam_val  is not None: return exam_val
        else:
            if eval_avg is not None and exam_val is not None:
                return (eval_avg + exam_val) / 2
            if eval_avg is not None: return eval_avg
            if exam_val  is not None: return exam_val
        return None

    # ---------- تصميم الخلية ----------
    def style_cell(self, cell, remark: str, lang: str,
                   grade=None, max_grade=None):
        if grade is not None and max_grade:
            hex_col = self.calc.get_grade_hex_color(grade, max_grade)
        else:
            color_map = self.config["COLORS"].get(
                lang, self.config["COLORS"]["ar"])
            hex_col = color_map.get(remark, "7F8C8D")

        cell.fill = PatternFill(fill_type="solid", fgColor=hex_col)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        thin = Side(style="thin", color="FFFFFF")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---------- المعالجة الرئيسية ----------
    def process_workbook(self, file_bytes):
        wb = openpyxl.load_workbook(file_bytes)
        report = []
        sheet_grade_map = {}   # {sheet_name: {row_idx: final_grade}}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if not ws.max_row or ws.max_row <= 1: continue

            subj, lang = self.detect_subject(sheet_name)
            cols = self.map_columns(ws, subj, self.settings["max_grade"])
            if not cols["found"]:
                self.logger.warning(
                    f"تخطي '{sheet_name}': لم يُعثر على أعمدة.")
                report.append({"الورقة": sheet_name,
                                "الحالة": "⏭️ متخطاة",
                                "ملاحظات أضيفت": 0,
                                "تم تخطيها": 0})
                continue

            sheet_max = self.calc.detect_max_grade(
                cols["exam_header"], self.settings["max_grade"])
            added, skipped = 0, 0
            sheet_grades = []
            student_grades_map = {}

            for row_idx in range(cols["header_row"] + 1,
                                 ws.max_row + 1):
                remark_cell = ws.cell(row=row_idx,
                                      column=cols["remark_col"])
                existing = remark_cell.value
                has_remark = (existing is not None
                              and str(existing).strip() != "")

                if self.settings["overwrite_mode"] == "تخطي الكل":
                    skipped += 1; continue
                if (self.settings["overwrite_mode"] == "ملء الفارغة فقط"
                        and has_remark):
                    skipped += 1; continue

                eval_vals, exam_val = [], None
                is_absent, is_exempt = False, False

                for c in cols["eval_cols"]:
                    p = self.calc.parse_grade(
                        ws.cell(row=row_idx, column=c).value)
                    if p["status"] == "absent": is_absent = True
                    elif p["status"] == "exempt": is_exempt = True
                    elif p["status"] == "valid":
                        eval_vals.append(p["value"])

                if cols["exam_col"]:
                    p = self.calc.parse_grade(
                        ws.cell(row=row_idx,
                                column=cols["exam_col"]).value)
                    if p["status"] == "absent": is_absent = True
                    elif p["status"] == "exempt": is_exempt = True
                    elif p["status"] == "valid": exam_val = p["value"]

                if is_absent or is_exempt:
                    txt = ("Absent" if lang == "fr" else
                           "Exempt" if lang == "en" else "غائب")
                    if is_exempt:
                        txt = ("Dispensé" if lang == "fr" else
                               "Exempt" if lang == "en" else "معفى")
                    remark_cell.value = txt
                    if self.settings["apply_style"]:
                        self.style_cell(remark_cell, txt, lang)
                    added += 1
                    continue

                weights = self.settings.get("custom_weights", {}).get(
                    subj, None)
                final = self.compute_final_grade(eval_vals, exam_val,
                                                  subj, weights)
                if final is not None:
                    capped = max(0.0, min(final, sheet_max))
                    sheet_grades.append(capped)
                    student_grades_map[row_idx] = capped
                    fr  = round(capped, self.settings["round_dec"])
                    txt = self.calc.get_remark(fr, sheet_max, lang)
                    remark_cell.value = txt
                    if self.settings["apply_style"]:
                        self.style_cell(remark_cell, txt, lang,
                                        grade=capped,
                                        max_grade=sheet_max)
                    added += 1
                else:
                    skipped += 1

            sheet_grade_map[sheet_name] = student_grades_map

            if sheet_grades:
                st_obj = self.calc.calc_statistics(sheet_grades, sheet_max)
                st_obj["sheet"] = sheet_name
                st_obj["subject"] = subj
                st_obj["lang"]    = lang
                st_obj["max"]     = sheet_max
                self.stats.append(st_obj)

                self.stats[-1].update({
                    "المادة": sheet_name,
                    "السلم": sheet_max,
                    "أعلى معدل": st_obj["max"],
                    "أدنى معدل": st_obj["min"],
                    "المعدل العام": st_obj["mean"],
                    "الوسيط": st_obj["median"],
                    "الانحراف المعياري": st_obj["std"],
                    "نسبة النجاح %": st_obj["pass_rate"],
                    "R1 (25%)": st_obj["q1"],
                    "R3 (75%)": st_obj["q3"],
                })

            report.append({
                "الورقة": sheet_name,
                "الحالة": f"✅ ({cols.get('notes', 'تلقائي')})",
                "ملاحظات أضيفت": added,
                "تم تخطيها": skipped
            })

        # ====== إضافة لوحة إحصائيات مُحسَّنة ======
        self.generate_dashboard(wb)

        # ====== ترتيب التلاميذ عبر المواد ======
        self.generate_ranking_sheet(wb, sheet_grade_map, wb.sheetnames)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, report

    # ---------- لوحة الإحصائيات ----------
    def generate_dashboard(self, wb):
        if not self.stats: return
        ws = wb.create_sheet(title="📊 لوحة الإحصائيات")
        ws.sheet_view.rightToLeft = True

        headers = ["المادة", "السلم", "أعلى معدل", "أدنى معدل",
                   "المعدل العام", "الوسيط", "الانحراف المعياري",
                   "نسبة النجاح %", "R1 (25%)", "R3 (75%)"]

        col_widths = [20, 8, 12, 12, 14, 10, 18, 16, 10, 10]
        for i, (h, w) in enumerate(zip(headers, col_widths), 1):
            ws.column_dimensions[get_column_letter(i)].width = w
            c = ws.cell(row=1, column=i)
            c.value = h
            c.font  = Font(bold=True, color="FFFFFF", size=11)
            c.fill  = PatternFill(fill_type="solid", fgColor="1A252F")
            c.alignment = Alignment(horizontal="center",
                                    vertical="center")
            ws.row_dimensions[1].height = 30

        for r, stat in enumerate(self.stats, 2):
            values = [
                stat.get("المادة", ""),
                stat.get("السلم", 10),
                stat.get("أعلى معدل", 0),
                stat.get("أدنى معدل", 0),
                stat.get("المعدل العام", 0),
                stat.get("الوسيط", 0),
                stat.get("الانحراف المعياري", 0),
                stat.get("نسبة النجاح %", 0),
                stat.get("R1 (25%)", 0),
                stat.get("R3 (75%)", 0),
            ]
            for ci, val in enumerate(values, 1):
                c = ws.cell(row=r, column=ci)
                c.value = val
                c.alignment = Alignment(horizontal="center",
                                        vertical="center")
                bg = "F2F3F4" if r % 2 == 0 else "FDFEFE"
                c.fill = PatternFill(fill_type="solid", fgColor=bg)
                thin = Side(style="thin", color="BDC3C7")
                c.border = Border(left=thin, right=thin,
                                  top=thin, bottom=thin)
                # تلوين نسبة النجاح
                if ci == 8 and isinstance(val, (int, float)):
                    hex_c = ("1A6B3C" if val >= 85 else
                             "27AE60" if val >= 70 else
                             "F39C12" if val >= 50 else "E74C3C")
                    c.fill = PatternFill(fill_type="solid",
                                        fgColor=hex_c)
                    c.font = Font(bold=True, color="FFFFFF")
            ws.row_dimensions[r].height = 22

        # رسم بياني عمودي
        chart = BarChart()
        chart.type, chart.style, chart.grouping = "col", 10, "clustered"
        chart.title = "المعدل العام ونسبة النجاح بالمواد"
        chart.y_axis.title, chart.x_axis.title = "النقطة", "المادة"
        chart.shape = 4
        nr = len(self.stats) + 1

        data_avg  = Reference(ws, min_col=5, max_col=5,
                               min_row=1, max_row=nr)
        data_pass = Reference(ws, min_col=8, max_col=8,
                               min_row=1, max_row=nr)
        cats = Reference(ws, min_col=1, min_row=2, max_row=nr)
        chart.add_data(data_avg,  titles_from_data=True)
        chart.add_data(data_pass, titles_from_data=True)
        chart.set_categories(cats)
        chart.width, chart.height = 22, 14
        ws.add_chart(chart, "A" + str(len(self.stats) + 4))

    # ---------- ورقة الترتيب ----------
    def generate_ranking_sheet(self, wb, grade_map, sheet_names):
        """يحسب معدلات التلاميذ عبر جميع المواد ويرتبهم"""
        if not grade_map: return

        # جمع جميع صفوف التلاميذ
        row_scores = {}
        for sheet, student_map in grade_map.items():
            subj, _ = self.detect_subject(sheet)
            weight = self.config["SUBJECT_PATTERNS"].get(
                subj, {}).get("weight", 1)
            for row_idx, grade in student_map.items():
                if row_idx not in row_scores:
                    row_scores[row_idx] = {"sum": 0, "weight": 0,
                                           "grades": {}}
                row_scores[row_idx]["sum"]    += grade * weight
                row_scores[row_idx]["weight"] += weight
                row_scores[row_idx]["grades"][sheet] = grade

        if not row_scores: return

        ranked = []
        for row_idx, data in row_scores.items():
            if data["weight"] > 0:
                gen = data["sum"] / data["weight"]
                ranked.append((row_idx, round(gen, 2), data["grades"]))
        ranked.sort(key=lambda x: x[1], reverse=True)

        ws = wb.create_sheet(title="🏆 الترتيب العام")
        ws.sheet_view.rightToLeft = True

        # رأس
        ws.merge_cells("A1:F1")
        ws["A1"].value = "🏆 الترتيب العام للتلاميذ"
        ws["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill  = PatternFill(fill_type="solid", fgColor="1A252F")
        ws["A1"].alignment = Alignment(horizontal="center",
                                       vertical="center")
        ws.row_dimensions[1].height = 32

        cols_h = ["الرتبة", "رقم الصف", "المعدل العام المُرجَّح",
                  "التقدير", "نسبة التحصيل %"]
        for ci, h in enumerate(cols_h, 1):
            c = ws.cell(row=2, column=ci)
            c.value = h
            c.font  = Font(bold=True, color="FFFFFF")
            c.fill  = PatternFill(fill_type="solid", fgColor="2980B9")
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(ci)].width = 18
        ws.row_dimensions[2].height = 24

        for rank, (row_idx, gen, _) in enumerate(ranked, 1):
            r = rank + 2
            remark = self.calc.get_remark(
                gen, self.settings["max_grade"], "ar")
            hex_c  = self.calc.get_grade_hex_color(
                gen, self.settings["max_grade"])
            pct    = round(gen / self.settings["max_grade"] * 100, 1)

            vals = [TextHelper.arabic_ordinal(rank), row_idx,
                    gen, remark, pct]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=ci)
                c.value = v
                c.alignment = Alignment(horizontal="center",
                                        vertical="center")
                bg = "F8F9FA" if r % 2 == 0 else "FFFFFF"
                c.fill = PatternFill(fill_type="solid", fgColor=bg)
                if ci in [3, 4, 5]:
                    c.fill = PatternFill(fill_type="solid", fgColor=hex_c)
                    c.font = Font(bold=True, color="FFFFFF")
            ws.row_dimensions[r].height = 20

        # تمييز المراتب الثلاث الأولى
        medals = ["🥇", "🥈", "🥉"]
        for mi in range(min(3, len(ranked))):
            c = ws.cell(row=mi + 3, column=1)
            c.value = medals[mi]
            c.font  = Font(size=14)
            c.fill  = PatternFill(fill_type="solid",
                                  fgColor=["FFD700", "C0C0C0",
                                           "CD7F32"][mi])


# =====================================================================
# 5. رسومات Plotly التحليلية
# =====================================================================

class DashboardCharts:

    @staticmethod
    def radar_chart(stats_list: list) -> go.Figure:
        cats = [s["المادة"] for s in stats_list]
        vals = [s["نسبة النجاح %"] for s in stats_list]
        vals_avg = [s["المعدل العام"] / s["السلم"] * 100
                    for s in stats_list]

        fig = go.Figure()
        fig.add_trace(go.Scatterpolar(
            r=vals + [vals[0]], theta=cats + [cats[0]],
            fill="toself", name="نسبة النجاح %",
            line_color="#27AE60", fillcolor="rgba(39,174,96,0.25)"))
        fig.add_trace(go.Scatterpolar(
            r=vals_avg + [vals_avg[0]], theta=cats + [cats[0]],
            fill="toself", name="المعدل (مُطبَّق على 100)",
            line_color="#2980B9", fillcolor="rgba(41,128,185,0.2)"))
        fig.update_layout(
            polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
            showlegend=True,
            title=dict(text="مخطط الرادار — أداء القسم", x=0.5),
            paper_bgcolor="#0E1117", plot_bgcolor="#0E1117",
            font=dict(color="white", family="Tajawal"),
            height=430
        )
        return fig

    @staticmethod
    def histogram_chart(grades: list, subject: str,
                         max_grade: float) -> go.Figure:
        fig = go.Figure()
        fig.add_trace(go.Histogram(
            x=grades, nbinsx=10,
            marker=dict(
                color=grades,
                colorscale=[
                    [0.0, "#C0392B"], [0.5, "#F39C12"], [1.0, "#27AE60"]
                ],
                cmin=0, cmax=max_grade,
                line=dict(color="white", width=0.5)
            ),
            name="توزيع النقاط",
            hovertemplate="النطاق: %{x}<br>عدد التلاميذ: %{y}<extra></extra>"
        ))
        fig.update_layout(
            title=f"توزيع نقاط — {subject}",
            xaxis_title="النقطة", yaxis_title="عدد التلاميذ",
            bargap=0.08,
            paper_bgcolor="#0E1117", plot_bgcolor="#161B22",
            font=dict(color="white", family="Tajawal"),
            height=340
        )
        return fig

    @staticmethod
    def bar_comparison(stats_list: list) -> go.Figure:
        subjects  = [s["المادة"] for s in stats_list]
        means     = [s["المعدل العام"] for s in stats_list]
        pass_rates = [s["نسبة النجاح %"] for s in stats_list]
        maxima    = [s["أعلى معدل"] for s in stats_list]
        minima    = [s["أدنى معدل"] for s in stats_list]

        fig = make_subplots(
            rows=1, cols=2,
            subplot_titles=("المعدل العام بالمواد",
                            "نسبة النجاح % بالمواد"),
            horizontal_spacing=0.12
        )
        fig.add_trace(go.Bar(
            x=subjects, y=means, name="المعدل العام",
            marker_color="#2980B9",
            error_y=dict(
                type="data",
                symmetric=False,
                array=[m - av for m, av in zip(maxima, means)],
                arrayminus=[av - mi for av, mi in zip(means, minima)],
                visible=True
            ),
            hovertemplate="%{x}<br>المعدل: %{y:.2f}<extra></extra>"
        ), row=1, col=1)

        colors = ["#27AE60" if p >= 70 else
                  "#F39C12" if p >= 50 else "#E74C3C"
                  for p in pass_rates]
        fig.add_trace(go.Bar(
            x=subjects, y=pass_rates, name="نسبة النجاح %",
            marker_color=colors,
            hovertemplate="%{x}<br>النجاح: %{y:.1f}%<extra></extra>"
        ), row=1, col=2)

        fig.update_layout(
            paper_bgcolor="#0E1117", plot_bgcolor="#161B22",
            font=dict(color="white", family="Tajawal"),
            showlegend=False, height=380
        )
        return fig

    @staticmethod
    def donut_distribution(distribution: dict, subject: str) -> go.Figure:
        labels = list(distribution.keys())
        values = list(distribution.values())
        colors = ["#1A6B3C", "#27AE60", "#52BE80", "#F1C40F",
                  "#E67E22", "#E74C3C", "#C0392B"]
        fig = go.Figure(go.Pie(
            labels=labels, values=values,
            hole=0.55,
            marker_colors=colors,
            textinfo="label+percent",
            hovertemplate="%{label}: %{value} تلميذ<extra></extra>"
        ))
        fig.update_layout(
            title=f"توزيع التقديرات — {subject}",
            annotations=[dict(text=subject, x=0.5, y=0.5,
                              font_size=12, showarrow=False,
                              font_color="white")],
            paper_bgcolor="#0E1117",
            font=dict(color="white", family="Tajawal"),
            height=360
        )
        return fig

    @staticmethod
    def gauge_pass_rate(rate: float, subject: str) -> go.Figure:
        color = ("#27AE60" if rate >= 70 else
                 "#F39C12" if rate >= 50 else "#E74C3C")
        fig = go.Figure(go.Indicator(
            mode="gauge+number+delta",
            value=rate,
            delta={"reference": 70, "valueformat": ".1f"},
            gauge={
                "axis": {"range": [0, 100], "tickcolor": "white"},
                "bar":  {"color": color},
                "steps": [
                    {"range": [0, 50],   "color": "#2C3E50"},
                    {"range": [50, 70],  "color": "#1A252F"},
                    {"range": [70, 100], "color": "#1E3A2F"},
                ],
                "threshold": {
                    "line": {"color": "gold", "width": 3},
                    "thickness": 0.85, "value": 70
                }
            },
            title={"text": f"نسبة النجاح<br>{subject}",
                   "font": {"color": "white", "size": 12}},
            number={"suffix": "%", "font": {"color": "white", "size": 28}}
        ))
        fig.update_layout(
            paper_bgcolor="#0E1117",
            font=dict(color="white", family="Tajawal"),
            height=280
        )
        return fig


# =====================================================================
# 6. الواجهة الرئيسية (Streamlit UI v2)
# =====================================================================

def inject_css():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@300;400;500;700;800&display=swap');

    html, body, [class*="css"], .stApp {
        font-family: 'Tajawal', sans-serif !important;
        direction: rtl;
        background: #0E1117;
        color: #ECF0F1;
    }

    /* ====== شريط العنوان الرئيسي ====== */
    .hero-banner {
        background: linear-gradient(135deg, #1A252F 0%, #2C3E50 50%, #1A6B3C 100%);
        border-radius: 16px;
        padding: 32px 24px;
        margin-bottom: 24px;
        text-align: center;
        border: 1px solid rgba(39,174,96,0.3);
        box-shadow: 0 8px 32px rgba(0,0,0,0.4);
        position: relative;
        overflow: hidden;
    }
    .hero-banner::before {
        content: '';
        position: absolute; inset: 0;
        background: radial-gradient(ellipse at top right,
            rgba(39,174,96,0.15) 0%, transparent 60%);
    }
    .hero-title {
        font-size: 2.4rem; font-weight: 800;
        color: #FFFFFF; margin: 0 0 8px 0;
        text-shadow: 0 2px 12px rgba(0,0,0,0.5);
    }
    .hero-subtitle {
        font-size: 1.05rem; color: #BDC3C7;
        margin: 0; font-weight: 300;
    }
    .hero-badge {
        display: inline-block;
        background: rgba(39,174,96,0.25);
        border: 1px solid #27AE60;
        border-radius: 20px;
        padding: 3px 14px;
        font-size: 0.82rem;
        color: #2ECC71;
        margin-top: 10px;
    }

    /* ====== بطاقات الإحصاء ====== */
    .kpi-card {
        background: linear-gradient(135deg, #1A252F, #2C3E50);
        border-radius: 12px;
        padding: 20px 16px;
        text-align: center;
        border: 1px solid rgba(255,255,255,0.08);
        box-shadow: 0 4px 16px rgba(0,0,0,0.3);
        transition: transform 0.2s, box-shadow 0.2s;
        min-height: 110px;
    }
    .kpi-card:hover {
        transform: translateY(-4px);
        box-shadow: 0 8px 24px rgba(0,0,0,0.5);
    }
    .kpi-value {
        font-size: 2.2rem; font-weight: 800;
        background: linear-gradient(90deg, #27AE60, #2980B9);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    .kpi-label {
        font-size: 0.88rem; color: #95A5A6;
        margin-top: 4px; font-weight: 500;
    }
    .kpi-icon { font-size: 1.5rem; margin-bottom: 4px; }

    /* ====== بطاقات المواد ====== */
    .subject-card {
        background: #1A252F;
        border-radius: 10px;
        padding: 14px 16px;
        border-left: 4px solid;
        margin-bottom: 10px;
        display: flex;
        justify-content: space-between;
        align-items: center;
    }
    .subject-name { font-weight: 700; font-size: 1rem; }
    .subject-avg  { font-size: 1.3rem; font-weight: 800; }

    /* ====== قسم الرفع ====== */
    .upload-zone {
        background: linear-gradient(135deg,
            rgba(26,37,47,0.8), rgba(44,62,80,0.8));
        border: 2px dashed rgba(41,128,185,0.5);
        border-radius: 16px;
        padding: 36px 24px;
        text-align: center;
        transition: border-color 0.3s;
    }
    .upload-zone:hover { border-color: #27AE60; }

    /* ====== الشريط الجانبي ====== */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0D1B2A 0%, #1A252F 100%);
        border-right: 1px solid rgba(255,255,255,0.06);
    }
    [data-testid="stSidebar"] * {
        direction: rtl !important;
        text-align: right !important;
        color: #ECF0F1 !important;
    }
    .sidebar-section {
        background: rgba(255,255,255,0.04);
        border-radius: 10px;
        padding: 12px;
        margin-bottom: 12px;
        border: 1px solid rgba(255,255,255,0.06);
    }

    /* ====== أزرار ====== */
    .stButton > button {
        width: 100%; border-radius: 10px;
        font-weight: 700; font-size: 1rem;
        background: linear-gradient(135deg, #27AE60, #1A6B3C);
        color: white; border: none;
        padding: 12px 20px;
        box-shadow: 0 4px 15px rgba(39,174,96,0.35);
        transition: all 0.3s;
    }
    .stButton > button:hover {
        background: linear-gradient(135deg, #2ECC71, #27AE60);
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(39,174,96,0.5);
    }

    /* ====== التبويبات ====== */
    .stTabs [data-baseweb="tab"] {
        font-family: 'Tajawal', sans-serif;
        font-weight: 600; font-size: 0.95rem;
        color: #95A5A6;
    }
    .stTabs [aria-selected="true"] {
        color: #27AE60 !important;
        border-bottom: 2px solid #27AE60;
    }

    /* ====== جداول البيانات ====== */
    .dataframe { direction: rtl; }
    .dataframe th {
        background: #1A252F !important;
        color: #ECF0F1 !important;
        font-weight: 700 !important;
        text-align: center !important;
    }

    /* ====== رسائل النجاح والخطأ ====== */
    .success-banner {
        background: linear-gradient(135deg,
            rgba(39,174,96,0.15), rgba(26,107,60,0.2));
        border: 1px solid rgba(39,174,96,0.4);
        border-radius: 12px;
        padding: 16px 20px;
        color: #2ECC71;
        font-weight: 600;
        text-align: center;
        font-size: 1.05rem;
    }

    /* ====== شريط التقدم ====== */
    .progress-bar-wrap {
        background: rgba(255,255,255,0.06);
        border-radius: 20px; overflow: hidden;
        height: 10px; margin: 6px 0;
    }
    .progress-bar-fill {
        height: 100%; border-radius: 20px;
        background: linear-gradient(90deg, #27AE60, #2980B9);
        transition: width 0.4s;
    }

    /* إخفاء عناصر Streamlit الافتراضية */
    #MainMenu, footer, header { visibility: hidden; }
    .block-container { padding-top: 1rem; }
    </style>
    """, unsafe_allow_html=True)


def kpi_card(icon, value, label, color="#27AE60"):
    return f"""
    <div class="kpi-card">
        <div class="kpi-icon">{icon}</div>
        <div class="kpi-value" style="
            background: linear-gradient(90deg, {color}, #2980B9);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;">
            {value}
        </div>
        <div class="kpi-label">{label}</div>
    </div>"""


def subject_progress_bar(name, avg, max_g, color):
    pct = min(int(avg / max_g * 100), 100) if max_g else 0
    return f"""
    <div class="subject-card" style="border-color:{color}">
        <div>
            <div class="subject-name" style="color:{color}">{name}</div>
            <div class="progress-bar-wrap" style="width:160px">
                <div class="progress-bar-fill"
                     style="width:{pct}%; background:{color}"></div>
            </div>
        </div>
        <div class="subject-avg" style="color:{color}">{avg}</div>
    </div>"""


# =====================================================================
# 7. نقطة الدخول الرئيسية
# =====================================================================

def run_ui():
    st.set_page_config(
        page_title="أداة الملاحظات الذكية v2",
        page_icon="🎓",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    inject_css()

    # -------- Hero Banner --------
    st.markdown("""
    <div class="hero-banner">
        <div class="hero-title">🎓 أداة الملاحظات الذكية</div>
        <div class="hero-subtitle">
            نظام متكامل لمعالجة ملفات الرقمنة الجزائرية —
            حساب آلي · ملاحظات ذكية · إحصائيات متقدمة · كارنيه رقمي
        </div>
        <div class="hero-badge">✦ الإصدار 2.0 — الطور الابتدائي ✦</div>
    </div>
    """, unsafe_allow_html=True)

    # ===================== الشريط الجانبي =====================
    with st.sidebar:
        st.markdown("## ⚙️ إعدادات المعالجة")

        # --- معلومات المدرسة ---
        with st.expander("🏫 معلومات المؤسسة", expanded=False):
            school_info = {
                "ministry": st.text_input(
                    "الوزارة", "وزارة التربية الوطنية"),
                "wilaya":   st.text_input("الولاية", ""),
                "school":   st.text_input("اسم المدرسة", ""),
                "class":    st.text_input("القسم", ""),
                "year":     st.text_input("السنة الدراسية",
                                          "2024 / 2025"),
                "trimester": st.selectbox(
                    "الفصل الدراسي",
                    MASTER_CONFIG["TRIMESTERS"])
            }

        st.divider()

        # --- إعدادات التقييم ---
        st.markdown("#### 📐 إعدادات التقييم")
        max_grade = st.selectbox(
            "سلم التنقيط الافتراضي:",
            options=[10, 20, 100], index=0)
        round_dec = st.selectbox(
            "دقة تقريب النقطة:",
            options=[2, 1, 0], index=0)

        st.divider()
        st.markdown("#### ⚖️ أوزان التقييمات المخصصة")
        use_custom_weights = st.checkbox(
            "تفعيل أوزان مخصصة؟", value=False)
        custom_weights = {}
        if use_custom_weights:
            for subj in ["arabic", "math", "french"]:
                lbl = MASTER_CONFIG["SUBJECT_PATTERNS"]\
                    .get(subj, {}).get("label", subj)
                c1, c2 = st.columns(2)
                with c1:
                    ew = st.number_input(
                        f"تقويم {lbl}", 0.1, 5.0, 1.0, 0.1,
                        key=f"ew_{subj}")
                with c2:
                    xw = st.number_input(
                        f"اختبار {lbl}", 0.1, 5.0, 1.0, 0.1,
                        key=f"xw_{subj}")
                custom_weights[subj] = {"eval": ew, "exam": xw}

        st.divider()
        st.markdown("#### 🔧 إعدادات أخرى")
        overwrite_mode = st.radio(
            "وضع الملاحظات الموجودة:",
            ["استبدال الموجودة", "ملء الفارغة فقط", "تخطي الكل"])
        ignore_zero = st.checkbox(
            "تجاهل الأصفار في حساب التقويمات؟", value=False)
        apply_style = st.checkbox(
            "تلوين خلايا الملاحظات (تدرج ألوان)؟", value=True)
        gen_bulletin = st.checkbox(
            "إنشاء كارنيه التلاميذ (تجريبي)؟", value=False)

        st.divider()
        st.markdown("#### 🎨 نظام الملاحظات")
        remark_lang = st.selectbox(
            "لغة الملاحظات الافتراضية:",
            ["عربي (تلقائي)", "فرنسي", "إنجليزي"])

        st.divider()
        # معاينة ألوان التقديرات
        with st.expander("🎨 معاينة سلم التقديرات", expanded=False):
            grade_preview = [
                ("ممتاز +", "#1A6B3C", "≥ 95%"),
                ("ممتاز",   "#27AE60", "≥ 85%"),
                ("جيد جداً","#52BE80", "≥ 75%"),
                ("جيد",     "#F1C40F", "≥ 65%"),
                ("مقبول",   "#E67E22", "≥ 55%"),
                ("متوسط",   "#E74C3C", "≥ 50%"),
                ("دون المتوسط","#C0392B","< 50%"),
            ]
            for lbl, col, pct in grade_preview:
                st.markdown(
                    f'<div style="background:{col};color:white;'
                    f'border-radius:6px;padding:4px 10px;'
                    f'margin:3px 0;text-align:center;font-weight:700;">'
                    f'{lbl} — {pct}</div>',
                    unsafe_allow_html=True)

    # ===================== المحتوى الرئيسي =====================

    user_settings = {
        "max_grade": max_grade,
        "round_dec": round_dec,
        "overwrite_mode": overwrite_mode,
        "ignore_zero": ignore_zero,
        "apply_style": apply_style,
        "custom_weights": custom_weights,
        "gen_bulletin": gen_bulletin,
    }

    # التبويبات الرئيسية
    tab_upload, tab_results, tab_analytics, tab_help = st.tabs([
        "📂 رفع الملف",
        "📋 نتائج المعالجة",
        "📊 التحليلات البيانية",
        "❓ المساعدة"
    ])

    # ===================== تبويب الرفع =====================
    with tab_upload:
        st.markdown("### 📂 رفع ملف الرقمنة")

        col_up, col_info = st.columns([3, 2])
        with col_up:
            uploaded_file = st.file_uploader(
                "اسحب وأفلت ملف الإكسيل هنا (.xlsx)",
                type=["xlsx"],
                help="يقبل النظام ملفات .xlsx فقط (ملفات الرقمنة الجزائرية)"
            )
            if uploaded_file:
                fsize = len(uploaded_file.getvalue()) / 1024
                st.markdown(
                    f'<div class="success-banner">'
                    f'✅ تم رفع الملف: <b>{uploaded_file.name}</b>'
                    f' ({fsize:.1f} KB)</div>',
                    unsafe_allow_html=True)

        with col_info:
            st.markdown("""
            <div style="background:#1A252F;border-radius:12px;
                        padding:16px;border:1px solid rgba(255,255,255,0.08)">
            <h4 style="color:#27AE60;margin-top:0">📌 تعليمات سريعة</h4>
            <ul style="color:#BDC3C7;font-size:0.9rem;padding-right:16px">
                <li>كل ورقة = مادة دراسية واحدة</li>
                <li>اسم الورقة يحدد المادة تلقائياً</li>
                <li>النقاط يجب أن تكون أرقاماً صحيحة</li>
                <li>استخدم (غ) للغياب و(م) للإعفاء</li>
                <li>عمود الملاحظات يُنشأ تلقائياً إن لم يُوجد</li>
            </ul>
            </div>
            """, unsafe_allow_html=True)

        if uploaded_file and st.button("🚀 بدء المعالجة الذكية",
                                        use_container_width=True):

            logger = logging.getLogger("GradeTool")
            logger.setLevel(logging.WARNING)
            log_stream = io.StringIO()
            ch = logging.StreamHandler(log_stream)
            ch.setFormatter(logging.Formatter("⚠️ %(message)s"))
            logger.handlers.clear()
            logger.addHandler(ch)

            progress_bar = st.progress(0, text="جاري التحضير...")
            status_txt   = st.empty()

            with st.spinner("جاري قراءة وتحليل ملف الرقمنة..."):
                progress_bar.progress(15, "قراءة الملف...")
                processor = ExcelProcessorV2(
                    MASTER_CONFIG, user_settings, logger)
                try:
                    progress_bar.progress(35, "معالجة الأوراق...")
                    processed_bytes, report_data = \
                        processor.process_workbook(uploaded_file)

                    # إنشاء كارنيه (تجريبي)
                    bulletin_bytes = None
                    if gen_bulletin and processor.stats:
                        progress_bar.progress(65, "إنشاء الكارنيه...")
                        bgen = BulletinGenerator(MASTER_CONFIG,
                                                  user_settings)
                        bwb  = openpyxl.Workbook()
                        bwb.remove(bwb.active)
                        sample_student = {
                            "name": "نموذج تلميذ",
                            "class": school_info.get("class", ""),
                            "subjects": {
                                s["subject"]: {
                                    "eval_avg": s["mean"] * 0.9,
                                    "exam":     s["mean"] * 1.1,
                                    "final":    s["mean"],
                                    "max":      s["max"]
                                }
                                for s in processor.stats
                                if "subject" in s
                            }
                        }
                        bgen.create_bulletin_sheet(
                            bwb, sample_student, school_info,
                            school_info.get("trimester",
                                            "الفصل الأول"))
                        bio = io.BytesIO()
                        bwb.save(bio); bio.seek(0)
                        bulletin_bytes = bio

                    progress_bar.progress(90, "حفظ الملف...")
                    st.session_state["processed_bytes"] = processed_bytes
                    st.session_state["report_data"]     = report_data
                    st.session_state["stats"]           = processor.stats
                    st.session_state["bulletin_bytes"]  = bulletin_bytes
                    st.session_state["file_name"]       = uploaded_file.name
                    progress_bar.progress(100, "✅ اكتمل!")

                    # -------- KPI Cards --------
                    total_added = sum(
                        r.get("ملاحظات أضيفت", 0) for r in report_data)
                    total_skip  = sum(
                        r.get("تم تخطيها", 0) for r in report_data)
                    avg_pass = (
                        round(sum(s.get("نسبة النجاح %", 0)
                                  for s in processor.stats) /
                              len(processor.stats), 1)
                        if processor.stats else 0
                    )
                    overall_avg = (
                        round(sum(s.get("المعدل العام", 0)
                                  for s in processor.stats) /
                              len(processor.stats), 2)
                        if processor.stats else 0
                    )

                    st.markdown("---")
                    k1, k2, k3, k4 = st.columns(4)
                    with k1: st.markdown(
                        kpi_card("✅", total_added,
                                 "ملاحظة أُضيفت", "#27AE60"),
                        unsafe_allow_html=True)
                    with k2: st.markdown(
                        kpi_card("📚", len(processor.stats),
                                 "مادة مُعالجة", "#2980B9"),
                        unsafe_allow_html=True)
                    with k3: st.markdown(
                        kpi_card("📈", f"{avg_pass}%",
                                 "معدل النجاح العام",
                                 "#F39C12" if avg_pass < 70 else "#27AE60"),
                        unsafe_allow_html=True)
                    with k4: st.markdown(
                        kpi_card("🏆", overall_avg,
                                 "المعدل الكلي للقسم", "#9B59B6"),
                        unsafe_allow_html=True)

                    st.markdown("---")

                    # -------- أزرار التحميل --------
                    dl1, dl2 = st.columns(2)
                    with dl1:
                        st.download_button(
                            label="📥 تحميل ملف الرقمنة الجاهز",
                            data=processed_bytes,
                            file_name=uploaded_file.name.replace(
                                ".xlsx", "_جاهز.xlsx"),
                            mime="application/vnd.openxmlformats-"
                                 "officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    if bulletin_bytes:
                        with dl2:
                            st.download_button(
                                label="📄 تحميل الكارنيه (نموذج)",
                                data=bulletin_bytes,
                                file_name="كارنيه_نموذج.xlsx",
                                mime="application/vnd.openxmlformats-"
                                     "officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )

                    logs = log_stream.getvalue()
                    if logs:
                        with st.expander("⚠️ سجل التحذيرات"):
                            st.code(logs)

                except Exception as e:
                    st.error(f"❌ خطأ غير متوقع: {e}")
                    import traceback
                    st.code(traceback.format_exc())

    # ===================== تبويب النتائج =====================
    with tab_results:
        if "report_data" not in st.session_state:
            st.info("⬆️ ارفع ملفاً وشغّل المعالجة لعرض النتائج.")
        else:
            st.markdown("### 📋 تقرير المعالجة التفصيلي")
            report_df = pd.DataFrame(st.session_state["report_data"])
            st.dataframe(report_df, use_container_width=True,
                         hide_index=True)

            if st.session_state.get("stats"):
                st.markdown("### 📊 جدول الإحصائيات الكاملة")
                stats_df = pd.DataFrame([
                    {k: v for k, v in s.items()
                     if k in ["المادة", "السلم", "أعلى معدل",
                               "أدنى معدل", "المعدل العام",
                               "الوسيط", "الانحراف المعياري",
                               "نسبة النجاح %"]}
                    for s in st.session_state["stats"]
                ])
                st.dataframe(stats_df, use_container_width=True,
                             hide_index=True)

                st.markdown("### 📌 أداء المواد (شريط التقدم)")
                for s in st.session_state["stats"]:
                    subj_key = s.get("subject", "other")
                    cfg = MASTER_CONFIG["SUBJECT_PATTERNS"].get(
                        subj_key, {})
                    color = cfg.get("color", "#27AE60")
                    label = cfg.get("label", s["المادة"])
                    icon  = cfg.get("icon", "📘")
                    st.markdown(
                        subject_progress_bar(
                            f"{icon} {label}",
                            s["المعدل العام"],
                            s["السلم"], color),
                        unsafe_allow_html=True)

    # ===================== تبويب التحليلات =====================
    with tab_analytics:
        if "stats" not in st.session_state or \
                not st.session_state["stats"]:
            st.info("⬆️ ارفع ملفاً أولاً لعرض التحليلات.")
        else:
            stats = st.session_state["stats"]

            st.markdown("### 📊 لوحة التحليلات البيانية التفاعلية")

            # --- مخطط الرادار ---
            c1, c2 = st.columns([2, 1])
            with c1:
                if len(stats) >= 3:
                    st.plotly_chart(
                        DashboardCharts.radar_chart(stats),
                        use_container_width=True)
                else:
                    st.info("يحتاج مخطط الرادار إلى 3 مواد على الأقل.")
            with c2:
                for s in stats:
                    st.plotly_chart(
                        DashboardCharts.gauge_pass_rate(
                            s["نسبة النجاح %"], s["المادة"]),
                        use_container_width=True)

            st.divider()

            # --- مقارنة المواد ---
            st.plotly_chart(
                DashboardCharts.bar_comparison(stats),
                use_container_width=True)

            st.divider()

            # --- توزيع التقديرات لكل مادة ---
            st.markdown("### 🍩 توزيع التقديرات بالمواد")
            cols_donut = st.columns(
                min(len(stats), 3))
            for i, s in enumerate(stats):
                with cols_donut[i % len(cols_donut)]:
                    if "distribution" in s:
                        st.plotly_chart(
                            DashboardCharts.donut_distribution(
                                s["distribution"], s["المادة"]),
                            use_container_width=True)

            # --- جدول الإحصائيات الموسع ---
            st.divider()
            st.markdown("### 📐 إحصائيات المئينيات")
            pc_data = []
            for s in stats:
                pc_data.append({
                    "المادة": s["المادة"],
                    "R1 (25%)": s.get("R1 (25%)", "—"),
                    "الوسيط (50%)": s.get("الوسيط", "—"),
                    "R3 (75%)": s.get("R3 (75%)", "—"),
                    "المدى": round(
                        s.get("أعلى معدل", 0) -
                        s.get("أدنى معدل", 0), 2),
                    "σ (الانحراف)": s.get(
                        "الانحراف المعياري", "—"),
                })
            st.dataframe(pd.DataFrame(pc_data),
                         use_container_width=True, hide_index=True)

            # --- تصدير التقرير CSV ---
            st.divider()
            csv_df = pd.DataFrame([
                {k: v for k, v in s.items()
                 if not isinstance(v, dict)}
                for s in stats
            ])
            csv_bytes = csv_df.to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                "📤 تصدير الإحصائيات (CSV)",
                data=csv_bytes,
                file_name="إحصائيات_القسم.csv",
                mime="text/csv",
                use_container_width=True
            )

    # ===================== تبويب المساعدة =====================
    with tab_help:
        st.markdown("## ❓ دليل الاستخدام الكامل")

        with st.expander("📂 كيفية تحضير ملف الرقمنة", expanded=True):
            st.markdown("""
            <div dir="rtl" style="line-height:2">

            ### 1️⃣ هيكل الملف المطلوب:
            - كل ورقة (Sheet) تمثل **مادة دراسية واحدة**
            - **يجب أن يحتوي اسم الورقة على اسم المادة** مثل:
              `لغة عربية` أو `رياضيات` أو `français`

            ### 2️⃣ رؤوس الأعمدة:
            | العمود | المحتوى | مثال |
            |--------|---------|------|
            | الاسم / اللقب | بيانات التلميذ | محمد |
            | التقويم | نقطة التقويم المستمر | 7.5 |
            | الاختبار | نقطة الاختبار الفصلي | 8 |
            | الملاحظة | يُعبأ تلقائياً | جيد جداً |

            ### 3️⃣ رموز خاصة:
            - `غ` أو `غائب` → تلميذ غائب
            - `م` أو `معفى` → تلميذ معفى

            ### 4️⃣ السلم التلقائي:
            - إذا كان العنوان `/10` → السلم 10
            - إذا كان العنوان `/20` → السلم 20
            - وإلا يُستخدم السلم الافتراضي من الإعدادات

            </div>
            """, unsafe_allow_html=True)

        with st.expander("🎨 نظام التقديرات والألوان"):
            data_rem = {
                "النسبة": ["≥ 95%", "≥ 85%", "≥ 75%",
                            "≥ 65%", "≥ 55%", "≥ 50%", "< 50%"],
                "التقدير (عربي)": ["ممتاز +", "ممتاز", "جيد جداً",
                                    "جيد", "مقبول", "متوسط",
                                    "دون المتوسط"],
                "التقدير (فرنسي)": ["Excellent +", "Excellent",
                                     "Très bien", "Bien", "Passable",
                                     "Suffisant", "Insuffisant"],
                "اللون": ["🟢🟢", "🟢", "🟢", "🟡", "🟠", "🔴", "🔴"],
            }
            st.dataframe(pd.DataFrame(data_rem),
                         use_container_width=True, hide_index=True)

        with st.expander("⚙️ الأوزان والمعادلات"):
            st.markdown("""
            <div dir="rtl" style="line-height:2">

            **المعادلة الأساسية (مع أوزان متساوية):**
            ```
            المعدل النهائي = (متوسط التقويمات + نقطة الاختبار) ÷ 2
            ```

            **مع أوزان مخصصة (مثال: تقويم×1 + اختبار×2):**
            ```
            المعدل = (نقطة_التقويم × 1 + نقطة_الاختبار × 2) ÷ 3
            ```

            **المعدل العام المُرجَّح للقسم:**
            ```
            المعدل العام = Σ(معدل_المادة × وزن_المادة) ÷ Σ الأوزان
            ```
            أوزان المواد الافتراضية: عربي وريا = 3، فرنسي وإسلامية وعلوم = 2، باقي المواد = 1

            </div>
            """, unsafe_allow_html=True)

        with st.expander("🐛 مشاكل شائعة وحلولها"):
            issues = {
                "المشكلة": [
                    "لم تُعثر أعمدة النقاط",
                    "الملاحظات غير صحيحة",
                    "المعدل العام صفر",
                    "خطأ في قراءة الملف"
                ],
                "الحل": [
                    "تأكد من رؤوس الأعمدة وأنها باللغة العربية",
                    "راجع السلم المختار (10 أو 20)",
                    "تأكد أن الخلايا تحتوي أرقاماً لا نصاً",
                    "تأكد أن الملف بصيغة .xlsx وليس .xls"
                ]
            }
            st.dataframe(pd.DataFrame(issues),
                         use_container_width=True, hide_index=True)


# =====================================================================
if __name__ == "__main__":
    from streamlit import runtime
    if runtime.exists():
        run_ui()
    else:
        from streamlit.web import cli as stcli
        sys.argv = ["streamlit", "run", __file__]
        sys.exit(stcli.main())
